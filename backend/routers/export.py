"""
routers/export.py — Excel export endpoint.

GET /feedback/export
  Returns a downloadable .xlsx file containing ALL feedback columns
  including feedback_date and created_at (hidden in UI but visible in Excel).
"""

import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)

import models
from database import get_db

router = APIRouter(tags=["Export"])


# ─── Styling helpers ────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")   # light blue
THIN_BORDER   = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin"),
)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center")


COLUMNS = [
    ("ID",            "id",             10),
    ("Canteen Name",  "canteen_name",   20),
    ("Meal Type",     "meal_type",      14),
    ("Food Quality",  "food_quality",   14),
    ("Food Taste",    "food_taste",     12),
    ("Food Hygiene",  "food_hygiene",   14),
    ("Staff Behavior","staff_behavior", 16),
    ("Cleanliness",   "cleanliness",    13),
    ("Comments",      "comments",       40),
    ("Feedback Date", "feedback_date",  16),
    ("Submitted At",  "created_at",     22),
]

# Rating label map for Excel cells
RATING_LABEL = {1: "1 - Poor", 2: "2 - Fair", 3: "3 - Good", 4: "4 - Great", 5: "5 - Excellent"}


@router.get("/feedback/export")
def export_feedback_excel(
    meal_type:    Optional[str]  = Query(None),
    canteen_name: Optional[str]  = Query(None),
    from_date:    Optional[date] = Query(None),
    to_date:      Optional[date] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Export all feedback to a formatted Excel workbook.
    All columns are included — feedback_date and created_at are
    intentionally visible in Excel (unlike the web UI).
    """
    # ── Query ────────────────────────────────────────────────
    q = db.query(models.Feedback)
    if meal_type:
        q = q.filter(models.Feedback.meal_type == meal_type)
    if canteen_name:
        q = q.filter(models.Feedback.canteen_name == canteen_name)
    if from_date:
        q = q.filter(models.Feedback.feedback_date >= from_date)
    if to_date:
        q = q.filter(models.Feedback.feedback_date <= to_date)

    records = q.order_by(models.Feedback.feedback_date.desc()).all()

    # ── Workbook ─────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Canteen Feedback"

    # Title row
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value     = f"Canteen Feedback Report — Exported {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    title_cell.font      = Font(bold=True, size=13, color="1F4E79")
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 28

    # Header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, record in enumerate(records, start=3):
        is_alt = (row_idx % 2 == 0)
        fill   = ALT_FILL if is_alt else PatternFill()

        values = [
            record.id,
            record.canteen_name or "Main Canteen",
            record.meal_type,   # plain text: Breakfast / Lunch / Dinner
            RATING_LABEL.get(record.food_quality,   str(record.food_quality)),
            RATING_LABEL.get(record.food_taste,     str(record.food_taste)),
            RATING_LABEL.get(record.food_hygiene,   str(record.food_hygiene)),
            RATING_LABEL.get(record.staff_behavior, str(record.staff_behavior)),
            RATING_LABEL.get(record.cleanliness,    str(record.cleanliness)),
            (record.comments or "").encode('ascii', 'replace').decode('ascii')
                if record.comments else "",
            record.feedback_date.strftime("%d-%m-%Y")    if record.feedback_date else "",
            record.created_at.strftime("%d-%m-%Y %H:%M") if record.created_at    else "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = THIN_BORDER
            cell.fill      = fill
            cell.alignment = CENTER_ALIGN if col_idx != 9 else LEFT_ALIGN
        ws.row_dimensions[row_idx].height = 18

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Auto-filter
    ws.auto_filter.ref = f"A2:K{max(2, len(records) + 2)}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Export Summary"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A2"] = "Total Records"
    ws2["B2"] = len(records)
    ws2["A3"] = "Exported On"
    ws2["B3"] = datetime.now().strftime("%d %b %Y %H:%M")
    if meal_type:
        ws2["A4"] = "Meal Filter"
        ws2["B4"] = meal_type
    if canteen_name:
        ws2["A5"] = "Canteen Filter"
        ws2["B5"] = canteen_name

    # ── Stream response ──────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.read()

    filename = f"canteen_feedback_export_{date.today().strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length":      str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition, Content-Length",
        },
    )
